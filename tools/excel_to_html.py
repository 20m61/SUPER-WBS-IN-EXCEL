#!/usr/bin/env python3
"""
ExcelファイルをHTML形式でプレビュー生成する
"""
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import html
import sys


def cell_to_html_style(cell) -> str:
    """セルのスタイルをCSSに変換"""
    styles = []

    # フォント
    if cell.font:
        if cell.font.bold:
            styles.append("font-weight: bold")
        if cell.font.color and cell.font.color.rgb:
            color = cell.font.color.rgb
            if isinstance(color, str) and len(color) >= 6:
                styles.append(f"color: #{color[-6:]}")

    # 背景色
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        color = cell.fill.fgColor.rgb
        if isinstance(color, str) and len(color) >= 6 and color != "00000000":
            styles.append(f"background-color: #{color[-6:]}")

    # 配置
    if cell.alignment:
        if cell.alignment.horizontal:
            styles.append(f"text-align: {cell.alignment.horizontal}")

    return "; ".join(styles)


def sheet_to_html(sheet, max_rows=30, max_cols=15) -> str:
    """シートをHTMLテーブルに変換"""
    html_parts = []
    html_parts.append(f'<h2>{html.escape(sheet.title)}</h2>')
    html_parts.append('<table class="excel-table">')

    # ヘッダー行（列番号）
    html_parts.append('<tr><th></th>')
    for col in range(1, min(sheet.max_column + 1, max_cols + 1)):
        html_parts.append(f'<th>{get_column_letter(col)}</th>')
    html_parts.append('</tr>')

    # データ行
    for row in range(1, min(sheet.max_row + 1, max_rows + 1)):
        html_parts.append(f'<tr><th>{row}</th>')
        for col in range(1, min(sheet.max_column + 1, max_cols + 1)):
            cell = sheet.cell(row=row, column=col)
            value = cell.value

            # 数式の場合は表示
            if isinstance(value, str) and value.startswith('='):
                display = f'<span class="formula">{html.escape(value[:50])}</span>'
            elif value is None:
                display = ''
            else:
                display = html.escape(str(value))

            style = cell_to_html_style(cell)
            style_attr = f' style="{style}"' if style else ''
            html_parts.append(f'<td{style_attr}>{display}</td>')
        html_parts.append('</tr>')

    html_parts.append('</table>')

    if sheet.max_row > max_rows:
        html_parts.append(f'<p class="note">... 他 {sheet.max_row - max_rows} 行</p>')

    return '\n'.join(html_parts)


def excel_to_html(xlsx_path: Path, output_path: Path, sheets: list = None):
    """ExcelファイルをHTMLに変換"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    html_content = '''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <title>Excel Preview</title>
    <style>
        body {
            font-family: "Meiryo UI", "Yu Gothic", sans-serif;
            background: #1a1a2e;
            color: #eee;
            padding: 20px;
            margin: 0;
        }
        h1 {
            color: #4a9eff;
            border-bottom: 2px solid #4a9eff;
            padding-bottom: 10px;
        }
        h2 {
            color: #8bc34a;
            margin-top: 30px;
            padding: 10px;
            background: #2c3e50;
            border-radius: 5px;
        }
        .excel-table {
            border-collapse: collapse;
            margin: 10px 0;
            font-size: 12px;
            width: 100%;
            table-layout: fixed;
        }
        .excel-table th, .excel-table td {
            border: 1px solid #444;
            padding: 6px 8px;
            text-align: left;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            max-width: 200px;
        }
        .excel-table th {
            background: #2c3e50;
            color: #ecf0f1;
            font-weight: bold;
            text-align: center;
        }
        .excel-table tr:nth-child(even) {
            background: #16213e;
        }
        .excel-table tr:hover {
            background: #1f4068;
        }
        .formula {
            color: #ff9800;
            font-family: monospace;
            font-size: 10px;
        }
        .note {
            color: #888;
            font-style: italic;
        }
        .sheet-nav {
            background: #2c3e50;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 20px;
        }
        .sheet-nav a {
            color: #4a9eff;
            margin-right: 15px;
            text-decoration: none;
        }
        .sheet-nav a:hover {
            text-decoration: underline;
        }
        .summary {
            background: #16213e;
            padding: 15px;
            border-radius: 5px;
            margin: 20px 0;
        }
        .summary h3 {
            color: #4a9eff;
            margin-top: 0;
        }
        .summary ul {
            list-style: none;
            padding: 0;
        }
        .summary li {
            padding: 5px 0;
            border-bottom: 1px solid #333;
        }
        .summary li:last-child {
            border-bottom: none;
        }
    </style>
</head>
<body>
'''

    # タイトル
    html_content += f'<h1>📊 {html.escape(xlsx_path.name)}</h1>\n'

    # シート一覧
    html_content += '<div class="sheet-nav">\n'
    html_content += '<strong>シート:</strong> '
    for name in wb.sheetnames:
        html_content += f'<a href="#{html.escape(name)}">{html.escape(name)}</a> '
    html_content += '\n</div>\n'

    # サマリー
    html_content += '<div class="summary">\n'
    html_content += '<h3>📋 ファイル情報</h3>\n'
    html_content += '<ul>\n'
    html_content += f'<li>シート数: {len(wb.sheetnames)}</li>\n'

    formula_count = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
    html_content += f'<li>数式数: {formula_count}</li>\n'
    html_content += '</ul>\n</div>\n'

    # 各シート
    sheet_names = sheets or wb.sheetnames
    for name in sheet_names:
        if name in wb.sheetnames:
            sheet = wb[name]
            html_content += f'<div id="{html.escape(name)}">\n'
            html_content += sheet_to_html(sheet)
            html_content += '</div>\n'

    html_content += '</body>\n</html>'

    output_path.write_text(html_content, encoding='utf-8')
    print(f"Generated: {output_path}")
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: excel_to_html.py <xlsx_file> [output.html]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else xlsx_path.with_suffix('.html')

    excel_to_html(xlsx_path, output_path)


if __name__ == "__main__":
    main()
