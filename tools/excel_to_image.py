#!/usr/bin/env python3
"""
ExcelファイルをPNG画像に変換する
Pillowを使用してブラウザ不要で画像を生成
"""
import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from pathlib import Path
import sys


# カラーパレット（Non-Excel Look ダークテーマ）
# 仕様書のUI/UX指針に基づく配色
COLORS = {
    'bg': (26, 26, 46),           # 背景
    'header_bg': (44, 62, 80),    # #2C3E50 ヘッダー背景
    'row_even': (22, 33, 62),     # #16213E 偶数行
    'row_odd': (26, 26, 46),      # 奇数行
    'input_bg': (234, 242, 248),  # #EAF2F8 入力セル背景
    'calc_bg': (245, 245, 245),   # #F5F5F5 計算セル背景
    'border': (68, 68, 68),       # 罫線
    'text': (238, 238, 238),      # テキスト
    'text_dark': (51, 51, 51),    # 暗い背景用テキスト
    'header_text': (255, 255, 255),
    'title': (74, 158, 255),      # #4A9EFF タイトル
    'formula': (255, 152, 0),     # #FF9800 数式
    'sheet_title': (139, 195, 74), # #8BC34A シートタイトル
    # ステータス色
    'status_todo': (189, 195, 199),    # #BDC3C7 未着手
    'status_doing': (52, 152, 219),    # #3498DB 進行中
    'status_delay': (231, 76, 60),     # #E74C3C 遅延
    'status_done': (46, 204, 113),     # #2ECC71 完了
}

# フォント設定
def get_font(size=12, bold=False):
    """フォントを取得（システムフォントにフォールバック）"""
    if bold:
        font_paths = [
            "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Bold.ttc",
            "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Medium.ttc",
            "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/liberation/LiberationSans-Bold.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        ]
    else:
        font_paths = [
            "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/google-noto-cjk/NotoSansCJK-Medium.ttc",
            "/usr/share/fonts/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        ]
    for path in font_paths:
        try:
            return ImageFont.truetype(path, size)
        except:
            continue
    return ImageFont.load_default()


def render_sheet(sheet, max_rows=25, max_cols=12):
    """シートを画像としてレンダリング"""
    # セルサイズ
    cell_width = 100
    cell_height = 25
    header_height = 30
    row_header_width = 40
    padding = 20
    title_height = 40

    # 画像サイズ計算
    cols = min(sheet.max_column, max_cols)
    rows = min(sheet.max_row, max_rows)

    img_width = row_header_width + cols * cell_width + padding * 2
    img_height = title_height + header_height + rows * cell_height + padding * 2

    # 画像作成
    img = Image.new('RGB', (img_width, img_height), COLORS['bg'])
    draw = ImageDraw.Draw(img)

    font = get_font(12)
    font_bold = get_font(12, bold=True)
    font_title = get_font(16, bold=True)
    font_small = get_font(10)

    # シートタイトル
    draw.text(
        (padding, padding),
        f"📊 {sheet.title}",
        fill=COLORS['sheet_title'],
        font=font_title
    )

    # 開始位置
    start_x = padding + row_header_width
    start_y = padding + title_height

    # 列ヘッダー
    for col in range(cols):
        x = start_x + col * cell_width
        draw.rectangle(
            [x, start_y, x + cell_width, start_y + header_height],
            fill=COLORS['header_bg'],
            outline=COLORS['border']
        )
        col_letter = get_column_letter(col + 1)
        text_bbox = draw.textbbox((0, 0), col_letter, font=font_bold)
        text_width = text_bbox[2] - text_bbox[0]
        draw.text(
            (x + (cell_width - text_width) // 2, start_y + 8),
            col_letter,
            fill=COLORS['header_text'],
            font=font_bold
        )

    # 行ヘッダーとデータ
    for row in range(rows):
        y = start_y + header_height + row * cell_height
        row_color = COLORS['row_even'] if row % 2 == 0 else COLORS['row_odd']

        # 行番号
        draw.rectangle(
            [padding, y, padding + row_header_width, y + cell_height],
            fill=COLORS['header_bg'],
            outline=COLORS['border']
        )
        row_num = str(row + 1)
        text_bbox = draw.textbbox((0, 0), row_num, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        draw.text(
            (padding + (row_header_width - text_width) // 2, y + 5),
            row_num,
            fill=COLORS['header_text'],
            font=font
        )

        # セルデータ
        for col in range(cols):
            x = start_x + col * cell_width
            cell = sheet.cell(row=row + 1, column=col + 1)

            # セル背景とテキスト色をスタイルから決定
            cell_color = row_color
            text_color_override = None

            # セルのスタイルから背景色を取得
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = cell.fill.fgColor.rgb
                if isinstance(rgb, str) and len(rgb) >= 6 and rgb != "00000000":
                    try:
                        r = int(rgb[-6:-4], 16)
                        g = int(rgb[-4:-2], 16)
                        b = int(rgb[-2:], 16)
                        cell_color = (r, g, b)
                        # 明るい背景色の場合は暗いテキストに
                        if r > 200 and g > 200 and b > 200:
                            text_color_override = COLORS['text_dark']
                        elif r > 150 or g > 150 or b > 150:
                            text_color_override = COLORS['text_dark']
                    except:
                        pass

            # ヘッダー行（4行目）の場合はヘッダースタイル
            if row + 1 == 4:
                cell_color = COLORS['header_bg']
                text_color_override = COLORS['header_text']

            draw.rectangle(
                [x, y, x + cell_width, y + cell_height],
                fill=cell_color,
                outline=COLORS['border']
            )

            # セル値
            value = cell.value
            if value is not None:
                if isinstance(value, str) and value.startswith('='):
                    # 数式は「計算値」として表示（実際のExcelでは計算結果が表示される）
                    # 数式の種類に応じてプレースホルダーを表示
                    if 'IFS(' in value or 'IF(' in value:
                        display = "―"  # 条件判定
                    elif 'WORKDAY(' in value or 'TODAY()' in value:
                        display = "日付"
                    elif 'SUM' in value or 'COUNT' in value or 'AVERAGE' in value:
                        display = "0"
                    elif 'FILTER(' in value or 'LET(' in value:
                        display = "（動的）"
                    elif 'INDIRECT(' in value:
                        display = "―"
                    elif 'HYPERLINK(' in value:
                        display = "🔗 リンク"
                    else:
                        display = "..."
                    text_color = COLORS['formula']
                    text_font = font_small
                elif isinstance(value, (int, float)):
                    # 数値
                    if 0 <= value <= 1 and value != int(value):
                        display = f"{value:.0%}"  # 進捗率
                    elif value > 40000 and value < 50000:
                        # Excelシリアル日付
                        display = "日付"
                    else:
                        display = str(int(value)) if value == int(value) else f"{value:.1f}"
                    text_color = COLORS['text']
                    text_font = font
                else:
                    display = str(value)[:15]
                    text_color = COLORS['text']
                    text_font = font

                # フォント色（オーバーライドがあればそれを使用）
                if text_color_override:
                    text_color = text_color_override
                elif cell.font and cell.font.color and cell.font.color.rgb:
                    rgb = cell.font.color.rgb
                    if isinstance(rgb, str) and len(rgb) >= 6:
                        try:
                            r = int(rgb[-6:-4], 16)
                            g = int(rgb[-4:-2], 16)
                            b = int(rgb[-2:], 16)
                            text_color = (r, g, b)
                        except:
                            pass

                draw.text(
                    (x + 4, y + 5),
                    display,
                    fill=text_color,
                    font=text_font
                )

    return img


def excel_to_images(xlsx_path: Path, output_dir: Path, sheets: list = None):
    """ExcelファイルをPNG画像群に変換"""
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    sheet_names = sheets or wb.sheetnames

    images = []
    for i, name in enumerate(sheet_names):
        if name in wb.sheetnames:
            sheet = wb[name]
            img = render_sheet(sheet)
            output_path = output_dir / f"{i+1:02d}_{name.replace('/', '_')}.png"
            img.save(output_path)
            images.append(output_path)
            print(f"Generated: {output_path}")

    # 全シートを結合した画像も生成
    if images:
        combined = create_combined_image(wb, sheet_names)
        combined_path = output_dir / "00_overview.png"
        combined.save(combined_path)
        print(f"Generated: {combined_path}")
        images.insert(0, combined_path)

    return images


def create_combined_image(wb, sheet_names):
    """複数シートを1つの画像にまとめる"""
    padding = 30
    title_height = 60

    # 各シートの画像を生成
    sheet_images = []
    for name in sheet_names[:4]:  # 最初の4シートのみ
        if name in wb.sheetnames:
            img = render_sheet(wb[name], max_rows=15, max_cols=10)
            sheet_images.append((name, img))

    if not sheet_images:
        return Image.new('RGB', (800, 600), COLORS['bg'])

    # 2x2グリッドで配置
    cols = 2
    rows = (len(sheet_images) + 1) // 2

    max_width = max(img.width for _, img in sheet_images)
    max_height = max(img.height for _, img in sheet_images)

    total_width = cols * max_width + (cols + 1) * padding
    total_height = title_height + rows * max_height + (rows + 1) * padding

    combined = Image.new('RGB', (total_width, total_height), COLORS['bg'])
    draw = ImageDraw.Draw(combined)

    # タイトル
    font_title = get_font(20, bold=True)
    draw.text(
        (padding, 15),
        "📊 Modern Excel PMS - Overview",
        fill=COLORS['title'],
        font=font_title
    )

    # 各シート画像を配置
    for i, (name, img) in enumerate(sheet_images):
        col = i % cols
        row = i // cols
        x = padding + col * (max_width + padding)
        y = title_height + padding + row * (max_height + padding)
        combined.paste(img, (x, y))

    return combined


def main():
    if len(sys.argv) < 2:
        print("Usage: excel_to_image.py <xlsx_file> [output_dir]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("output/screenshots")

    images = excel_to_images(xlsx_path, output_dir)
    print(f"\nGenerated {len(images)} images in {output_dir}")


if __name__ == "__main__":
    main()
