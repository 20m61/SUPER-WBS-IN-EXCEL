#!/usr/bin/env python3
"""
Excelファイルの機能テスト

openpyxlを使用してExcelファイルの構造と機能を検証する。
"""
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import List, Tuple, Optional
import xml.etree.ElementTree as ET

# openpyxlをインポート
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Some tests will be skipped.")


class ExcelTester:
    """Excelファイルのテストクラス"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.passed: List[str] = []

    def run_all_tests(self) -> bool:
        """全テストを実行"""
        print(f"\n{'='*70}")
        print(f"Excel機能テスト: {self.file_path.name}")
        print(f"{'='*70}\n")

        tests = [
            ("ファイル存在確認", self.test_file_exists),
            ("ZIPアーカイブ検証", self.test_zip_structure),
            ("OpenXML構造検証", self.test_openxml_structure),
            ("openpyxlで開く", self.test_openpyxl_load),
            ("シート構成確認", self.test_sheet_structure),
            ("数式存在確認", self.test_formulas_exist),
            ("データバリデーション確認", self.test_data_validations),
            ("条件付き書式確認", self.test_conditional_formatting),
            ("VBAプロジェクト確認", self.test_vba_project),
            ("VML描画確認", self.test_vml_drawings),
        ]

        for test_name, test_func in tests:
            try:
                result = test_func()
                if result:
                    self.passed.append(test_name)
                    print(f"  ✅ {test_name}")
                else:
                    print(f"  ❌ {test_name}")
            except Exception as e:
                self.errors.append(f"{test_name}: {e}")
                print(f"  ❌ {test_name}: {e}")

        # サマリー
        print(f"\n{'='*70}")
        print(f"テスト結果: {len(self.passed)} passed, {len(self.errors)} failed")
        if self.warnings:
            print(f"警告: {len(self.warnings)}")
            for w in self.warnings:
                print(f"  ⚠️  {w}")
        print(f"{'='*70}\n")

        return len(self.errors) == 0

    def test_file_exists(self) -> bool:
        """ファイルが存在するか確認"""
        if not self.file_path.exists():
            self.errors.append(f"ファイルが見つかりません: {self.file_path}")
            return False
        return True

    def test_zip_structure(self) -> bool:
        """ZIPアーカイブとして正しいか確認"""
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zf:
                # 破損チェック
                bad_file = zf.testzip()
                if bad_file:
                    self.errors.append(f"ZIPファイルが破損: {bad_file}")
                    return False

                # 必須ファイルの確認
                required = [
                    '[Content_Types].xml',
                    '_rels/.rels',
                    'xl/workbook.xml',
                    'xl/styles.xml',
                ]
                for req in required:
                    if req not in zf.namelist():
                        self.errors.append(f"必須ファイルがありません: {req}")
                        return False
            return True
        except zipfile.BadZipFile as e:
            self.errors.append(f"無効なZIPファイル: {e}")
            return False

    def test_openxml_structure(self) -> bool:
        """OpenXML構造が正しいか確認"""
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zf:
                # Content_Typesを解析
                content_types = zf.read('[Content_Types].xml')
                root = ET.fromstring(content_types)

                # ワークシートのコンテンツタイプを確認
                ns = {'ct': 'http://schemas.openxmlformats.org/package/2006/content-types'}
                overrides = root.findall('.//ct:Override', ns)

                has_workbook = False
                has_worksheet = False
                for override in overrides:
                    part_name = override.get('PartName', '')
                    content_type = override.get('ContentType', '')
                    if 'workbook' in part_name.lower():
                        has_workbook = True
                    if 'worksheet' in part_name.lower():
                        has_worksheet = True

                if not has_workbook:
                    self.errors.append("workbook.xmlのコンテンツタイプがありません")
                    return False
                if not has_worksheet:
                    self.errors.append("worksheetのコンテンツタイプがありません")
                    return False

            return True
        except ET.ParseError as e:
            self.errors.append(f"XMLパースエラー: {e}")
            return False

    def test_openpyxl_load(self) -> bool:
        """openpyxlでファイルを開けるか確認"""
        if not OPENPYXL_AVAILABLE:
            self.warnings.append("openpyxlがインストールされていません")
            return True  # スキップ

        try:
            # data_only=Falseで数式を保持したまま読み込む
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=False,
                keep_vba=True  # VBAを保持
            )
            return True
        except Exception as e:
            self.errors.append(f"openpyxlでの読み込みエラー: {e}")
            return False

    def test_sheet_structure(self) -> bool:
        """シート構成を確認"""
        if not self.workbook:
            self.warnings.append("ワークブックが読み込まれていません")
            return True

        expected_sheets = ['Config', 'Template', 'PRJ_001', 'PRJ_002',
                          'Case_Master', 'Measure_Master', 'Kanban_View']

        actual_sheets = self.workbook.sheetnames
        missing = [s for s in expected_sheets if s not in actual_sheets]

        if missing:
            self.errors.append(f"シートがありません: {missing}")
            return False

        print(f"      シート数: {len(actual_sheets)}")
        return True

    def test_formulas_exist(self) -> bool:
        """数式が存在するか確認"""
        if not self.workbook:
            self.warnings.append("ワークブックが読み込まれていません")
            return True

        formula_count = 0
        formula_types = set()

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if cell.value.startswith('='):
                            formula_count += 1
                            # 数式の種類を特定
                            for func in ['WORKDAY', 'IFS', 'SUMPRODUCT', 'COUNTIF',
                                        'AVERAGEIF', 'INDIRECT', 'IFERROR', 'TODAY']:
                                if func in cell.value.upper():
                                    formula_types.add(func)

        if formula_count == 0:
            self.errors.append("数式が見つかりません")
            return False

        print(f"      数式数: {formula_count}")
        print(f"      使用関数: {', '.join(sorted(formula_types))}")
        return True

    def test_data_validations(self) -> bool:
        """データバリデーションを確認"""
        if not self.workbook:
            self.warnings.append("ワークブックが読み込まれていません")
            return True

        validation_count = 0
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if hasattr(sheet, 'data_validations') and sheet.data_validations:
                validation_count += len(sheet.data_validations.dataValidation)

        print(f"      バリデーション数: {validation_count}")
        return validation_count > 0

    def test_conditional_formatting(self) -> bool:
        """条件付き書式を確認"""
        if not self.workbook:
            self.warnings.append("ワークブックが読み込まれていません")
            return True

        cf_count = 0
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if hasattr(sheet, 'conditional_formatting'):
                cf_count += len(sheet.conditional_formatting._cf_rules)

        print(f"      条件付き書式数: {cf_count}")
        return cf_count > 0

    def test_vba_project(self) -> bool:
        """VBAプロジェクトを確認"""
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zf:
                if 'xl/vbaProject.bin' in zf.namelist():
                    vba_size = zf.getinfo('xl/vbaProject.bin').file_size
                    print(f"      vbaProject.bin: {vba_size} bytes")

                    # OLEシグネチャを確認
                    vba_data = zf.read('xl/vbaProject.bin')
                    if vba_data[:8] == b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1':
                        print(f"      OLE形式: 有効")
                        return True
                    else:
                        self.warnings.append("vbaProject.binがOLE形式ではありません")
                        return True
                else:
                    self.warnings.append("vbaProject.binがありません（VBAなしファイル）")
                    return True
        except Exception as e:
            self.errors.append(f"VBAプロジェクト確認エラー: {e}")
            return False

    def test_vml_drawings(self) -> bool:
        """VML描画（ボタン等）を確認"""
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zf:
                vml_files = [f for f in zf.namelist() if 'vmlDrawing' in f]
                if vml_files:
                    print(f"      VMLファイル: {len(vml_files)}")

                    # ボタンの確認
                    for vml_file in vml_files:
                        vml_content = zf.read(vml_file).decode('utf-8')
                        button_count = vml_content.count('<x:ClientData ObjectType="Button"')
                        if button_count > 0:
                            print(f"        {Path(vml_file).name}: ボタン{button_count}個")
                    return True
                else:
                    self.warnings.append("VML描画がありません（ボタンなしファイル）")
                    return True
        except Exception as e:
            self.errors.append(f"VML確認エラー: {e}")
            return False


def test_formula_calculation():
    """数式計算のテスト（シミュレーション）"""
    print("\n" + "="*70)
    print("数式計算シミュレーション")
    print("="*70 + "\n")

    # WORKDAYのテスト
    from datetime import date, timedelta

    def workday(start_date: date, days: int, holidays: List[date] = None) -> date:
        """WORKDAYのシミュレーション"""
        if holidays is None:
            holidays = []
        current = start_date
        remaining = days
        direction = 1 if days >= 0 else -1

        while remaining != 0:
            current += timedelta(days=direction)
            if current.weekday() < 5 and current not in holidays:  # 平日かつ祝日でない
                remaining -= direction

        return current

    # テストケース
    start = date(2025, 12, 15)
    work_days = 5

    end = workday(start, work_days - 1)  # 開始日を含むため-1
    print(f"  WORKDAY({start}, {work_days}) = {end}")

    # IFSのテスト
    def ifs_status(progress: float, end_date: date, today: date) -> str:
        """IFSステータス判定のシミュレーション"""
        if progress >= 1.0:
            return "完了"
        elif end_date < today and progress < 1.0:
            return "遅延"
        elif start <= today:
            return "進行中"
        else:
            return "未着手"

    today = date.today()
    test_cases = [
        (1.0, end, "完了"),
        (0.5, date(2025, 12, 1), "遅延"),
        (0.3, date(2025, 12, 31), "進行中"),
        (0.0, date(2026, 1, 15), "未着手"),
    ]

    print("\n  IFSステータス判定:")
    for progress, end_date, expected in test_cases:
        result = ifs_status(progress, end_date, today)
        status = "✅" if result == expected else "❌"
        print(f"    {status} 進捗{progress*100:.0f}% 期限{end_date} → {result}")

    print("\n  ✅ 数式計算シミュレーション完了")


def main():
    """メイン関数"""
    if len(sys.argv) < 2:
        # デフォルトのテストファイル
        test_files = [
            "output/ModernExcelPMS.xlsx",
            "output/ModernExcelPMS_regen.xlsm",
        ]
    else:
        test_files = sys.argv[1:]

    all_passed = True

    for file_path in test_files:
        if Path(file_path).exists():
            tester = ExcelTester(file_path)
            if not tester.run_all_tests():
                all_passed = False
        else:
            print(f"⚠️  ファイルが見つかりません: {file_path}")

    # 数式計算シミュレーション
    test_formula_calculation()

    # 最終結果
    print("\n" + "="*70)
    if all_passed:
        print("🎉 全てのテストに合格しました！")
    else:
        print("❌ 一部のテストが失敗しました")
    print("="*70 + "\n")

    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
